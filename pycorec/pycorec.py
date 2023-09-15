import csv  # system included
import ctypes  # system included
import ctypes.wintypes  # system included
import datetime  # system included
from itertools import zip_longest  # system included
from pathlib import Path  # system included
import re  # system included
from typing import Callable  # system included
from typing import Union  # system included
import customtkinter  # pip install customtkinter
from natsort import natsorted  # pip install natsort
from openpyxl.styles import Font  # pip install openpyxl
from openpyxl import Workbook  # pip install openpyxl
from PIL import Image, ImageTk  # pip install pillow
from screeninfo import get_monitors  # pip install screeninfo

pycorec_version = "2.0.4"


class PyCorec:
    def __init__(self):
        super().__init__()
        self.first_run = True
        self.magnification = 1
        self.photo_image = None
        self.resized_image = None
        self.image_file_name = None
        self.image_height = 1
        self.image_width = 1
        self.resized_image_height = 1
        self.resized_image_width = 1
        self.cm_per_px_x = None
        self.cm_per_px_y = None
        self.fps = None
        self.interval = 1
        self.image_paths = []
        self.current_image_index = 0
        self.coordinates = []
        self.pos = []
        self.file_list = []
        self.zoom_level = 1.0
        self.offset_x = 0
        self.offset_y = 0

        # create window
        self.root = customtkinter.CTk()
        self.root.title(f"PyCorec {pycorec_version}")
        customtkinter.set_appearance_mode("Dark")

        # configure window size
        def get_dpi_scale():
            # Create a temporary window to get DPI scale
            tmp_root = ctypes.windll.user32.CreateWindowExW(0, "STATIC", None,
                                                            0, 0, 0, 0, 0, 0, 0, 0, None)
            dpi_x = ctypes.windll.user32.GetDpiForWindow(tmp_root)
            ctypes.windll.user32.DestroyWindow(tmp_root)
            return dpi_x / 96.0  # 96 DPI is the default DPI

        def get_taskbar_height():
            hwnd = ctypes.windll.user32.FindWindowW("Shell_traywnd", None)
            rect = ctypes.wintypes.RECT()
            ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
            return int(rect.bottom) - int(rect.top)

        screen_size = get_monitors()
        primary_monitor = None
        screen_width = 0
        screen_height = 0
        for monitor in screen_size:
            if monitor.is_primary:
                primary_monitor = monitor
                break

        if primary_monitor:
            dpi_scale = get_dpi_scale()
            lower_adjustment = 0.5  # exclude taskbar:1.7, fullscreen: 0.5
            taskbar_height = get_taskbar_height()
            screen_width = int(primary_monitor.width / dpi_scale)
            screen_height = int((primary_monitor.height - taskbar_height * lower_adjustment) / dpi_scale)

        self.root.geometry(f"{screen_width}x{screen_height}+0+0")
        self.root.attributes('-topmost', True)

        # frames
        self.frame = customtkinter.CTkFrame(self.root)
        self.frame.pack(fill=customtkinter.BOTH, expand=True)

        self.image_frame = customtkinter.CTkFrame(self.frame)
        self.image_frame.pack(side=customtkinter.LEFT, fill=customtkinter.BOTH, expand=True)

        self.canvas = customtkinter.CTkCanvas(self.image_frame, bg="white")
        self.canvas.pack(fill=customtkinter.BOTH, expand=True)

        self.bottom_frame = customtkinter.CTkFrame(self.root)
        self.bottom_frame.pack(side=customtkinter.BOTTOM, fill=customtkinter.X)

        self.label_frame = customtkinter.CTkFrame(self.bottom_frame)
        self.label_frame.pack(side=customtkinter.LEFT, fill=customtkinter.X, padx=10, pady=10)

        # labels
        self.coordinates_label = customtkinter.CTkLabel(self.label_frame, text="Coordinates (px): (0, 0)")
        self.coordinates_label.pack(side=customtkinter.LEFT)

        self.records_label = customtkinter.CTkLabel(self.label_frame, text="Record Points: 0")
        self.records_label.pack(side=customtkinter.LEFT, padx=10)

        self.frame_interval_label = customtkinter.CTkLabel(self.label_frame, text="Frame Interval: 1")
        self.frame_interval_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_size_label = customtkinter.CTkLabel(self.label_frame, text="Image Size: ")
        self.image_size_label.pack(side=customtkinter.LEFT, padx=10)

        self.resized_image_size_label = customtkinter.CTkLabel(self.label_frame, text="Displayed Image Size: ")
        self.resized_image_size_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_magnification_label = customtkinter.CTkLabel(self.label_frame, text="Image Magnification (%): ")
        self.image_magnification_label.pack(side=customtkinter.LEFT, padx=10)

        self.fps_label = customtkinter.CTkLabel(self.label_frame, text="fps: ")
        self.fps_label.pack(side=customtkinter.LEFT, padx=10)

        self.cm_per_px_x_label = customtkinter.CTkLabel(self.label_frame, text="cm/px (x): ")
        self.cm_per_px_x_label.pack(side=customtkinter.LEFT, padx=10)

        self.cm_per_px_y_label = customtkinter.CTkLabel(self.label_frame, text="cm/px (y): ")
        self.cm_per_px_y_label.pack(side=customtkinter.LEFT, padx=10)

        #  buttons
        self.button_frame = customtkinter.CTkFrame(self.frame)
        self.button_frame.pack(side=customtkinter.RIGHT, fill=customtkinter.Y, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Open Images by",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_dir_button = customtkinter.CTkButton(self.button_frame, text="Directory Selection",
                                                             command=self.get_dir)
        self.open_image_dir_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_range_button = customtkinter.CTkButton(self.button_frame, text="Bounded Selection",
                                                               command=self.get_range)
        self.open_image_range_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_click_button = customtkinter.CTkButton(self.button_frame, text="Click Selection",
                                                               command=self.get_files)
        self.open_image_click_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent",
                                                    hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Move Image", fg_color="transparent",
                                                    hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=0)

        self.move_button = ArrowButton(self.button_frame, command=self.move_image)
        self.move_button.pack(fill=customtkinter.X, padx=70, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Image Magnification (%)",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.zoom_spinbox = FloatSpinbox(self.button_frame, width=150, step_size=3)
        self.zoom_spinbox.pack(fill=customtkinter.X, padx=10, pady=0)

        self.zoom_in_button = customtkinter.CTkButton(self.button_frame, text="Apply", command=self.zoom_image)
        self.zoom_in_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.fit_image_to_window_button = customtkinter.CTkButton(self.button_frame, text="Reset to Window Size",
                                                                  command=self.fit_image_to_window)
        self.fit_image_to_window_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.fit_image_to_actual_size_button = customtkinter.CTkButton(self.button_frame, text="Actual Size",
                                                                       command=self.fit_image_to_actual_size)
        self.fit_image_to_actual_size_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.save_file_button = customtkinter.CTkButton(self.button_frame, text="Save as...",
                                                        command=self.save_file)
        self.save_file_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Usage\n"
                                                                            "Next Image : → right arrow key\n"
                                                                            "Previous Image : ← left arrow key\n"
                                                                            "Record Coordinates : left-click\n"
                                                                            "Delete Coordinates : right-click",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        #  canvas
        self.canvas.bind("<Button-1>", self.on_canvas_left_click)
        self.canvas.bind("<Button-3>", self.on_canvas_right_click)
        self.canvas.bind("<Button-2>", self.on_canvas_wheel_click)
        self.canvas.bind("<Motion>", self.on_canvas_motion)
        self.canvas.configure(cursor="crosshair")

        self.root.bind("<Right>", self.next_image_keyboard)
        self.root.bind("<Left>", self.previous_image_keyboard)

    def configure_optional_parameter(self):
        fps_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                       " fps (Optional):\n"
                                                       "\n"
                                                       "If you want to add a time column to the output file, "
                                                       "enter a value of fps.",
                                                  title="fps (Optional)")
        self.fps = fps_dialog.get_input()
        if self.fps != "":
            self.fps = float(self.fps)
        self.fps_label.configure(text=f"fps: {self.fps}")

        cm_per_px_x_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                               "x cm/px (Optional):\n"
                                                               "\n"
                                                               "If you want to add physical coordinates (cm) converted "
                                                               "from image coordinates (px) to the output file, "
                                                               "enter a value of x-axis direction cm/px.",
                                                          title="x cm/px (Optional)")
        self.cm_per_px_x = cm_per_px_x_dialog.get_input()
        if self.cm_per_px_x != "":
            self.cm_per_px_x = float(self.cm_per_px_x)
        self.cm_per_px_x_label.configure(text=f"cm/px (x): {self.cm_per_px_x}")

        cm_per_px_y_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                               "y cm/px (Optional):\n"
                                                               "\n"
                                                               "If you want to add physical coordinates (cm) converted "
                                                               "from image coordinates (px) to the output file, "
                                                               "enter a value of y-axis direction cm/px.",
                                                          title="y cm/px (Optional)")
        self.cm_per_px_y = cm_per_px_y_dialog.get_input()
        if self.cm_per_px_y != "":
            self.cm_per_px_y = float(self.cm_per_px_y)
        self.cm_per_px_y_label.configure(text=f"cm/px (y): {self.cm_per_px_y}")

    def get_dir(self):
        dir_path = customtkinter.filedialog.askdirectory(title='Open Images from Directory ( jpg, png, tif, bmp )')
        path = Path(dir_path)
        image_paths = natsorted([p for p in path.glob('**/*')
                                 if not re.search('Bkg', str(p))
                                 and re.search('/*\.(jpg|jpeg|png|tif|bmp)', str(p))])
        if len(image_paths) != 0:
            interval_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                                "Frame Interval:\n"
                                                                "\n"
                                                                "Examples\n"
                                                                "1: Load all frames (001.jpg, 002.jpg, 003.jpg, ...)\n"
                                                                "2: Load frames every one frame (001,003,005,...)\n"
                                                                "3: Load frames every two frames (001,004,007,...)\n",
                                                           title="Frame Interval")
            self.interval = int(interval_dialog.get_input())
            self.frame_interval_label.configure(text=f"Frame Interval: {self.interval}")
            self.image_paths = image_paths[::self.interval]
            self.configure_optional_parameter()
            self.load_image()

    def get_range(self):
        file_type = [("Supported Files", "*.jpg *.JPG *.jpeg *.png *.PNG *.bmp *.BMP *.tif")]
        image_paths = customtkinter.filedialog.askopenfilenames(title="Open Images by Bounded Selection",
                                                                filetypes=file_type)
        image_paths = natsorted(image_paths)
        if len(image_paths) != 0:
            interval_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                                "Frame Interval:\n"
                                                                "\n"
                                                                "Examples\n"
                                                                "1: Load all frames (001.jpg, 002.jpg, 003.jpg, ...)\n"
                                                                "2: Load frames every one frame (001,003,005,...)\n"
                                                                "3: Load frames every two frames (001,004,007,...)\n",
                                                           title="Frame Interval")
            self.interval = int(interval_dialog.get_input())
            self.frame_interval_label.configure(text=f"Frame Interval: {self.interval}")
            self.image_paths = image_paths[::self.interval]
            self.configure_optional_parameter()
            self.load_image()

    def get_files(self):
        file_type = [("Supported Files", "*.jpg *.JPG *.jpeg *.png *.PNG *.bmp *.BMP *.tif")]
        image_paths = customtkinter.filedialog.askopenfilenames(title="Open Image(s) by Click", filetypes=file_type)
        if len(image_paths) != 0:
            self.image_paths = natsorted(image_paths)
            self.configure_optional_parameter()
            self.load_image()

    def load_image(self, fit_image=False):
        image_path = self.image_paths[self.current_image_index]
        image = Image.open(image_path)
        self.image_file_name = Path(image_path).name
        self.image_width, self.image_height = image.size
        self.image_size_label.configure(text=f"Image Size: {self.image_width} x {self.image_height}")
        self.resized_image = self.resize_image(image, fit_image)
        self.photo_image = ImageTk.PhotoImage(self.resized_image)
        self.canvas.delete("all")
        self.canvas.create_image(0 + self.offset_x, 0 + self.offset_y, anchor=customtkinter.NW,
                                 image=self.photo_image)
        self.update_labels(image_path)

    def resize_image(self, image, fit_image=False):
        image_width, image_height = image.size
        win_width = self.image_frame.winfo_width()
        win_height = self.image_frame.winfo_height()
        win_fit_magnification_ratio = min(win_width / image_width, win_height / image_height)
        if self.first_run:
            self.zoom_level = win_fit_magnification_ratio
            self.first_run = False
        if fit_image:
            self.zoom_level = win_fit_magnification_ratio
        new_width = int(image_width * self.zoom_level)
        new_height = int(image_height * self.zoom_level)
        if new_width <= 0 or new_height <= 0:
            return image
        resized_image = image.resize((new_width, new_height), Image.LANCZOS)
        self.resized_image_width = resized_image.width
        self.resized_image_height = resized_image.height
        self.magnification = self.resized_image_width / self.image_width
        return resized_image

    def zoom_image(self):
        self.zoom_level = self.zoom_spinbox.get() * 0.01
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def fit_image_to_window(self):
        fit_image = True
        self.offset_x = 0
        self.offset_y = 0
        self.canvas.delete("all")
        self.load_image(fit_image)
        self.draw_coordinates()

    def fit_image_to_actual_size(self):
        self.zoom_level = 1.0
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def move_image(self, dx, dy):
        self.offset_x += dx
        self.offset_y += dy
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def next_image_keyboard(self, event):
        self.next_image()

    def previous_image_keyboard(self, event):
        self.previous_image()

    def on_canvas_motion(self, event):
        x = (event.x - self.offset_x) / self.magnification
        y = (event.y - self.offset_y) / self.magnification
        self.update_coordinates_label(round(x), round(y))

    def on_canvas_left_click(self, event):
        if self.image_file_name is not None:
            x = (event.x - self.offset_x) / self.magnification
            y = (event.y - self.offset_y) / self.magnification
            self.coordinates.append((x, y))
            self.draw_coordinates()

    def on_canvas_right_click(self, event):
        if self.image_file_name is not None:
            if len(self.coordinates) != 0:
                del self.coordinates[-1]
                self.draw_coordinates()

    def on_canvas_wheel_click(self, event):
        if self.image_file_name is not None:
            x = float("nan")
            y = float("nan")
            self.coordinates.append((x, y))
            self.draw_coordinates()

    def draw_coordinates(self):
        self.canvas.delete("coordinates")
        for coord in self.coordinates:
            x, y = coord
            x = x * self.magnification + self.offset_x
            y = y * self.magnification + self.offset_y
            self.canvas.create_oval(x - 2, y - 2, x + 2, y + 2, fill="red", outline="red", tags="coordinates")
        record_points_number = len(self.coordinates)
        self.update_records_label(record_points_number)

    def record_coordinates(self):
        if len(self.pos) != self.current_image_index:
            self.pos[self.current_image_index] = self.coordinates
            self.file_list[self.current_image_index] = self.image_file_name
        if len(self.pos) == self.current_image_index:
            self.pos.insert(self.current_image_index, self.coordinates)
            self.file_list.insert(self.current_image_index, self.image_file_name)

    def next_image(self):
        if self.current_image_index + 1 == len(self.image_paths):
            self.save_file()
        if self.current_image_index + 1 < len(self.image_paths):
            self.record_coordinates()
            self.current_image_index = (self.current_image_index + 1) % len(self.image_paths)
            self.canvas.delete("all")
            self.load_image()
            if len(self.pos) != self.current_image_index:
                self.coordinates = self.pos[self.current_image_index]
                self.draw_coordinates()
            if len(self.pos) == self.current_image_index:
                self.coordinates = []
                self.draw_coordinates()

    def previous_image(self):
        if self.current_image_index != 0:
            if self.current_image_index + 1 == len(self.image_paths):
                self.record_coordinates()
            self.current_image_index = (self.current_image_index - 1) % len(self.image_paths)
            self.coordinates = self.pos[self.current_image_index]
            self.canvas.delete("all")
            self.load_image()
            self.draw_coordinates()

    def update_coordinates_label(self, x, y):
        self.coordinates_label.configure(text=f"Coordinates (px): ({x}, {y})")

    def update_records_label(self, record_points_number):
        self.records_label.configure(text=f"Record Points: {record_points_number}")

    def update_labels(self, image_path):
        self.root.title(f"PyCorec {pycorec_version}      [ {self.current_image_index + 1} / {len(self.image_paths)} ]  "
                        f"{Path(image_path).name}")
        self.resized_image_size_label.configure(text=f"Resized Image Size: "
                                                     f"{self.resized_image.width} x {self.resized_image.height}")
        image_magnification = (self.resized_image.width / self.image_width) * 100
        self.image_magnification_label.configure(text=f"Image Magnification (%): {image_magnification:.1f}")
        self.zoom_spinbox.set(image_magnification)

    def save_file(self):
        self.record_coordinates()
        now = datetime.datetime.now()
        current_time = now.strftime("%Y-%m-%d-%H-%M-%S")
        save_path = customtkinter.filedialog.asksaveasfilename(title="Save Coordinates File (Excel Book or CSV)",
                                                               filetypes=[("Excel Book", ".xlsx"),
                                                                          ("CSV", ".csv")],
                                                               initialfile=f"PyCorec"
                                                                           f"{pycorec_version}_{current_time}",
                                                               defaultextension=".xlsx")

        def write_list_1d_to_excel_column(worksheet, data_list_1d, start_row, column_number):
            for row_num, value in enumerate(data_list_1d, start=start_row):
                worksheet.cell(row=row_num, column=column_number, value=value)

        def write_list_2d_to_excel(worksheet, data_list_2d, start_row, start_col):
            for y, row in enumerate(data_list_2d):
                for x, cell_value in enumerate(row):
                    worksheet.cell(row=start_row + y,
                                   column=start_col + x,
                                   value=data_list_2d[y][x])

        if ".xlsx" in save_path:
            wb = Workbook()
            # Time Series by Points
            wb.create_sheet(index=0, title="Time Series by Points")
            ws_p = wb["Time Series by Points"]
            ws_p["A1"] = "Index"
            ws_p["B1"] = "Filename"
            for i in range(len(self.file_list)):
                ws_p.cell(row=i + 2, column=1, value=i)
                ws_p.cell(row=i + 2, column=2, value=self.file_list[i])

            if self.fps != "":
                ws_p["C1"] = "Time_s"
                spf = 1 / self.fps
                timestep = self.interval * spf
                end = 0 + (len(self.image_paths) - 1) * timestep
                sec_list = [0 + i * timestep for i in range(int((end - 0) / timestep) + 1)]
                for i in range(len(sec_list)):
                    ws_p.cell(row=i + 2, column=3, value=sec_list[i])

            flat_pos = [[item for sublist in row for item in sublist] for row in self.pos]
            flat_pos_row_length_check = [len(v) for v in flat_pos]
            flat_pos_row_length = max(flat_pos_row_length_check)
            flat_pos_row_set_length = flat_pos_row_length / 2

            for i in range(1, int(flat_pos_row_set_length + 1)):
                ws_p.cell(row=1, column=4 + 2 * (i - 1), value=f"x{i}_px")
                ws_p.cell(row=1, column=5 + 2 * (i - 1), value=f"y{i}_px")
            write_list_2d_to_excel(worksheet=ws_p, data_list_2d=flat_pos, start_row=2, start_col=4)

            if self.cm_per_px_x != "" and self.cm_per_px_y != "":
                cm_pos = [[d for d in row] for row in flat_pos]
                cm_pos = [[self.cm_per_px_x * value if column % 2 != 1 else value for column, value in enumerate(row)]
                          for row in cm_pos]
                cm_pos = [
                    [-1 * self.cm_per_px_y * value if column % 2 != 0 else value for column, value in enumerate(row)]
                    for row in cm_pos]
                cm_pos_row_length_check = [len(v) for v in cm_pos]
                cm_pos_row_length = max(cm_pos_row_length_check)
                cm_pos_row_set_length = cm_pos_row_length / 2

                for i in range(1, int(cm_pos_row_set_length + 1)):
                    ws_p.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (i - 1),
                              value=f"x{i}_cm")
                    ws_p.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 1 + 2 * (i - 1),
                              value=f"y{i}_cm")
                write_list_2d_to_excel(worksheet=ws_p, data_list_2d=cm_pos, start_row=2,
                                       start_col=4 + 2 * (flat_pos_row_set_length + 1 - 1))

                ws_p.cell(row=1,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1),
                          value=f"xm_px")
                ws_p.cell(row=2,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1),
                          value=self.image_width)
                ws_p.cell(row=1,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 1,
                          value=f"ym_px")
                ws_p.cell(row=2,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 1,
                          value=self.image_height)
                ws_p.cell(row=1,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 2,
                          value=f"xm_cm")
                ws_p.cell(row=2,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 2,
                          value=self.image_width * self.cm_per_px_x)
                ws_p.cell(row=1,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 3,
                          value=f"ym_cm")
                ws_p.cell(row=2,
                          column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 3,
                          value=self.image_height * self.cm_per_px_y * -1)

            font = Font(name="Segoe UI")
            for column in range(1, int(ws_p.max_column) + 1):
                for row in range(1, int(ws_p.max_row) + 1):
                    cell = ws_p.cell(row=row, column=column)
                    cell.font = font

            # Spatial Distribution per Frame
            wb.create_sheet(index=1, title="Spatial Distribution per Frame")
            ws_f = wb["Spatial Distribution per Frame"]
            ws_f["A1"] = "Index"
            ws_f["B1"] = "Filename"
            ws_f["D1"] = "PointIndex"
            ws_f["E1"] = "x_px"
            ws_f["F1"] = "y_px"

            tidy_pos = [list(tpl) for sublist in self.pos for tpl in sublist]
            record_point_counts = [len(row) for row in self.pos]
            index_col_list = []
            filename_col_list = []
            point_index_col_list = []
            for i in range(len(self.file_list)):
                for j in range(1, record_point_counts[i] + 1):
                    index_col_list.append(i)
                    filename_col_list.append(self.file_list[i])
                    point_index_col_list.append(j)

            write_list_1d_to_excel_column(worksheet=ws_f, data_list_1d=index_col_list, start_row=2, column_number=1)
            write_list_1d_to_excel_column(worksheet=ws_f, data_list_1d=filename_col_list, start_row=2, column_number=2)
            write_list_1d_to_excel_column(worksheet=ws_f, data_list_1d=point_index_col_list,
                                          start_row=2, column_number=4)
            write_list_2d_to_excel(worksheet=ws_f, data_list_2d=tidy_pos, start_row=2, start_col=5)

            if self.fps != "":
                ws_f["C1"] = "Time_s"
                spf = 1 / self.fps
                timestep = self.interval * spf
                end = 0 + (len(self.image_paths) - 1) * timestep
                sec_list = [0 + i * timestep for i in range(int((end - 0) / timestep) + 1)]
                sec_col_list = []
                for i in range(len(self.file_list)):
                    for j in range(1, record_point_counts[i] + 1):
                        sec_col_list.append(sec_list[i])
                write_list_1d_to_excel_column(worksheet=ws_f, data_list_1d=sec_col_list, start_row=2, column_number=3)

            if self.cm_per_px_x != "" and self.cm_per_px_y != "":
                ws_f["G1"] = "x_cm"
                ws_f["H1"] = "y_cm"
                tidy_cm_pos = [list(tpl) for sublist in self.pos for tpl in sublist]
                tidy_cm_pos = [[self.cm_per_px_x * value
                                if column == 0 else value for column, value in enumerate(row)]
                               for row in tidy_cm_pos]
                tidy_cm_pos = [
                    [-1 * self.cm_per_px_y * value if column == 1 else value for column, value in enumerate(row)]
                    for row in tidy_cm_pos]
                write_list_2d_to_excel(worksheet=ws_f, data_list_2d=tidy_cm_pos, start_row=2, start_col=7)

                ws_f["I1"] = "xm_px"
                ws_f.cell(row=2, column=9, value=self.image_width)
                ws_f["J1"] = "ym_px"
                ws_f.cell(row=2, column=10, value=self.image_height)
                ws_f["K1"] = "xm_cm"
                ws_f.cell(row=2, column=11, value=self.image_width * self.cm_per_px_x)
                ws_f["L1"] = "ym_cm"
                ws_f.cell(row=2, column=12, value=self.image_height * self.cm_per_px_y * -1)

            font = Font(name="Segoe UI")
            for column in range(1, int(ws_f.max_column) + 1):
                for row in range(1, int(ws_f.max_row) + 1):
                    cell = ws_f.cell(row=row, column=column)
                    cell.font = font

            wb.save(save_path)

        if ".csv" in save_path:
            col_names_list = ["Index", "Filename", "Time_s"]
            index_list = []
            sec_list = []
            pos_cm_list = []
            for i in range(len(self.file_list)):
                index_list.append(i)

            if self.fps != "":
                spf = 1 / self.fps
                timestep = self.interval * spf
                end = 0 + (len(self.image_paths) - 1) * timestep
                sec_list = [0 + i * timestep for i in range(int((end - 0) / timestep) + 1)]

            col_merge_list = list(zip_longest(index_list, self.file_list, sec_list, fillvalue=None))
            col_merge_list = [list(row) for row in col_merge_list]

            flat_pos = [[item for sublist in row for item in sublist] for row in self.pos]
            flat_pos_row_length_check = [len(v) for v in flat_pos]
            flat_pos_row_length = max(flat_pos_row_length_check)
            flat_pos_row_set_length = flat_pos_row_length / 2
            for i in range(int(flat_pos_row_set_length)):
                col_names_list.append(f"x{i + 1}_px")
                col_names_list.append(f"y{i + 1}_px")

            pos_merge_list = []
            for sublist1, sublist2 in zip(col_merge_list, flat_pos):
                sublist1.extend(sublist2)
                pos_merge_list.append(sublist1)
            pos_merge_list_length = max(len(v) for v in pos_merge_list)

            if self.cm_per_px_x != "" and self.cm_per_px_y != "":
                cm_pos = [[d for d in row] for row in flat_pos]
                cm_pos = [[self.cm_per_px_x * value if column % 2 != 1 else value for column, value in enumerate(row)]
                          for row in cm_pos]
                cm_pos = [
                    [-1 * self.cm_per_px_y * value if column % 2 != 0 else value for column, value in enumerate(row)]
                    for row in cm_pos]
                cm_pos_row_length_check = [len(v) for v in cm_pos]
                cm_pos_row_length = max(cm_pos_row_length_check)
                cm_pos_row_set_length = cm_pos_row_length / 2

                for i in range(int(cm_pos_row_set_length)):
                    col_names_list.append(f"x{i + 1}_cm")
                    col_names_list.append(f"y{i + 1}_cm")

                for row in pos_merge_list:
                    while len(row) < pos_merge_list_length:
                        row.append("")

                for row in cm_pos:
                    while len(row) < cm_pos_row_length:
                        row.append("")

                for sublist1, sublist2 in zip(pos_merge_list, cm_pos):
                    sublist1.extend(sublist2)
                    pos_cm_list.append(sublist1)

                col_names_list.extend([f"xm_px", f"ym_px", f"xm_cm", f"ym_cm"])
                image_size_list = [self.image_width, self.image_height,
                                   self.image_width * self.cm_per_px_x, self.image_height * self.cm_per_px_y * -1]
                for i in image_size_list:
                    pos_cm_list[0].append(i)

                pos_cm_list.insert(0, col_names_list)

                with open(save_path, "w", newline="") as file:
                    writer = csv.writer(file)
                    writer.writerows(pos_cm_list)

            if self.cm_per_px_x == "":
                pos_merge_list.insert(0, col_names_list)
                with open(save_path, "w", newline="") as file:
                    writer = csv.writer(file)
                    writer.writerows(pos_merge_list)


class FloatSpinbox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: Union[int, float] = 1,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        self.step_size = step_size
        self.command = command

        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure(2, weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=1)  # entry expands

        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height - 6, height=height - 6,
                                                       command=self.subtract_button_callback)
        self.subtract_button.grid(row=0, column=0, padx=(3, 0), pady=3)

        self.entry = customtkinter.CTkEntry(self, width=width - (2 * height), height=height - 6, border_width=0)
        self.entry.grid(row=0, column=1, columnspan=1, padx=3, pady=3, sticky="ew")

        self.add_button = customtkinter.CTkButton(self, text="+", width=height - 6, height=height - 6,
                                                  command=self.add_button_callback)
        self.add_button.grid(row=0, column=2, padx=(0, 3), pady=3)

        # default value
        self.entry.insert(0, "")

    def add_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = float(self.entry.get()) + self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def subtract_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = float(self.entry.get()) - self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def get(self) -> Union[float, None]:
        try:
            return float(self.entry.get())
        except ValueError:
            return None

    def set(self, value: float):
        self.entry.delete(0, "end")
        self.entry.insert(0, str(float(value)))


class ArrowButton(customtkinter.CTkFrame):
    def __init__(self, *args,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, **kwargs)

        self.command = command

        self.configure(fg_color="#333333")

        self.up_button = customtkinter.CTkButton(self, text="↑", height=20, width=20,
                                                 command=self.up_button_callback)
        self.up_button.grid(row=0, column=1, padx=1, pady=1)

        self.down_button = customtkinter.CTkButton(self, text="↓", height=20, width=20,
                                                   command=self.down_button_callback)
        self.down_button.grid(row=2, column=1, padx=1, pady=1)

        self.left_button = customtkinter.CTkButton(self, text="←", height=20, width=20,
                                                   command=self.left_button_callback)
        self.left_button.grid(row=1, column=0, padx=1, pady=1)

        self.down_button = customtkinter.CTkButton(self, text="→", height=20, width=20,
                                                   command=self.right_button_callback)
        self.down_button.grid(row=1, column=2, padx=1, pady=1)

    def up_button_callback(self):
        if self.command is not None:
            self.command(dx=0, dy=-10)
            return

    def down_button_callback(self):
        if self.command is not None:
            self.command(dx=0, dy=10)
            return

    def left_button_callback(self):
        if self.command is not None:
            self.command(dx=-10, dy=0)
            return

    def right_button_callback(self):
        if self.command is not None:
            self.command(dx=10, dy=0)
            return


if __name__ == "__main__":
    app = PyCorec()
    app.root.mainloop()
